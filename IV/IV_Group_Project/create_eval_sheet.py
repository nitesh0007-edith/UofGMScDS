"""Create the evaluation spreadsheet we hand to participants.

Four sheets: instructions, task reference, data collection grid, and
post-study ranking. Run this once to get a blank template.
"""

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join("evaluation", "Evaluation_Data_Collection.xlsx")

# --- Styling ---

HEADER_FILL = PatternFill(start_color="003865", end_color="003865", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
BODY_FONT = Font(name="Calibri", size=11)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
SYSTEM_FILLS = {
    "A": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "B": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    "C": PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
}


def _style_header_row(ws, row, num_cols):
    """Make the header row dark blue with white text."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_body_cell(cell, wrap=True):
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP_ALIGN if wrap else CENTER_ALIGN


# --- Sheet 1: Instructions ---

def _create_instructions(wb):
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = "003865"
    ws.column_dimensions["A"].width = 90

    instructions = [
        ("EVALUATION DATA COLLECTION — INSTRUCTIONS", True),
        ("", False),
        ("Study Design: Within-subjects (each participant uses all 3 systems)", False),
        ("", False),
        ("STEP 1: Assign participant order using Latin Square counterbalancing:", False),
        ("   P1, P4, P7 … → System A → B → C", False),
        ("   P2, P5, P8 … → System B → C → A", False),
        ("   P3, P6, P9 … → System C → A → B", False),
        ("", False),
        ("STEP 2: For EACH (participant × system × task) combination:", False),
        ("   a) Read the task aloud to the participant", False),
        ("   b) Start the timer when they begin interacting", False),
        ("   c) Stop the timer when they give their answer", False),
        ("   d) Record: time_seconds, accuracy (0/1/2), likert (1-5)", False),
        ("   e) Add any observations in the notes column", False),
        ("", False),
        ("STEP 3: After all tasks on all systems, fill the Post-Study sheet:", False),
        ("   Ask participant to rank systems 1 (best) to 3 (worst)", False),
        ("", False),
        ("SCORING GUIDE:", False),
        ("   accuracy = 0  →  Wrong answer or could not complete", False),
        ("   accuracy = 1  →  Partially correct (e.g. right direction but wrong value)", False),
        ("   accuracy = 2  →  Fully correct answer", False),
        ("", False),
        ("   likert = 1  →  Very difficult", False),
        ("   likert = 2  →  Difficult", False),
        ("   likert = 3  →  Neutral", False),
        ("   likert = 4  →  Easy", False),
        ("   likert = 5  →  Very easy", False),
        ("", False),
        ("IMPORTANT: Ensure participant explores the system independently.", False),
        ("Do NOT guide them toward the answer.", False),
    ]

    for i, (text, is_title) in enumerate(instructions, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if is_title:
            cell.font = Font(name="Calibri", bold=True, size=14, color="003865")
        else:
            cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(wrap_text=True)


# --- Sheet 2: Task Reference ---

TASKS = [
    {
        "id": "T1",
        "description": (
            "Identify the season or month with the highest average temperature range."
        ),
        "expected_answer": (
            "Season: Spring (avg tempRange = 7.77°C)\n"
            "Month: May (avg tempRange = 8.87°C)"
        ),
        "answer_format": "Name of season or month + numeric value",
        "hint": (
            "Use Metric dropdown → 'Avg Temp Range', "
            "Group by → 'By Season' or 'By Month'. "
            "Read the top bar."
        ),
    },
    {
        "id": "T2",
        "description": (
            "Find the top 3 weather types (desc) by frequency across the dataset."
        ),
        "expected_answer": (
            "1. rain (763 days)\n"
            "2. partly-cloudy-day (566 days)\n"
            "3. clear-day (210 days)"
        ),
        "answer_format": "3 weather type names in order, with counts",
        "hint": (
            "Use Metric dropdown → 'Count (Frequency)', "
            "Group by → 'Weather Type'. "
            "Read top 3 bars."
        ),
    },
    {
        "id": "T3",
        "description": "Determine which year had the highest average wind speed.",
        "expected_answer": "2019 (avg wind speed = 8.91 km/h)",
        "answer_format": "Year + numeric value",
        "hint": (
            "Use Metric dropdown → 'Avg Wind Speed', "
            "Group by → 'By Year'. "
            "Read top bar."
        ),
    },
    {
        "id": "T4",
        "description": (
            "Find the weather type with the highest average visibility, "
            "filtered by a minimum humidity threshold (humidity > 0.85)."
        ),
        "expected_answer": "clear-day (avg visibility = 5.61 km)",
        "answer_format": "Weather type name + numeric value",
        "hint": (
            "Set humidity slider > 0.85, "
            "Metric → 'Avg Visibility', "
            "Group by → 'Weather Type'. "
            "Read top bar."
        ),
    },
    {
        "id": "T5",
        "description": (
            "Identify months where wind speed is above a user-set threshold "
            "AND humidity is above a user-set threshold. "
            "Find the top 3 seasons with the highest average cloud cover."
        ),
        "expected_answer": (
            "Top 3 seasons by avg cloud cover:\n"
            "1. Winter (0.640)\n"
            "2. Autumn (0.594)\n"
            "3. Summer (0.591)"
        ),
        "answer_format": (
            "List of qualifying months (varies by threshold) + "
            "3 season names with cloud cover values"
        ),
        "hint": (
            "Set wind & humidity sliders to desired thresholds. "
            "Metric → 'Avg Cloud Cover', "
            "Group by → 'By Season'. "
            "Read top 3 bars."
        ),
    },
]


def _create_task_reference(wb):
    ws = wb.create_sheet("Task Reference")
    ws.sheet_properties.tabColor = "E74C3C"

    headers = [
        "Task ID", "Task Description", "Expected Answer",
        "Answer Format", "Hint (for evaluator only)",
    ]
    widths = [10, 45, 35, 30, 40]

    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    for row_idx, task in enumerate(TASKS, start=2):
        values = [
            task["id"],
            task["description"],
            task["expected_answer"],
            task["answer_format"],
            task["hint"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _style_body_cell(cell)
        ws.row_dimensions[row_idx].height = 60


# --- Sheet 3: Data Collection ---

PARTICIPANTS = ["P1", "P2", "P3", "P4", "P5"]
SYSTEMS = ["A", "B", "C"]
TASK_IDS = ["T1", "T2", "T3", "T4", "T5"]

# Latin square order
SYSTEM_ORDERS = {
    "P1": ["A", "B", "C"],
    "P2": ["B", "C", "A"],
    "P3": ["C", "A", "B"],
    "P4": ["A", "B", "C"],
    "P5": ["B", "C", "A"],
}


def _create_data_collection(wb):
    ws = wb.create_sheet("Data Collection")
    ws.sheet_properties.tabColor = "27AE60"

    headers = [
        "Participant", "System Order", "System", "Task",
        "Task Description (short)",
        "Time (seconds)", "Accuracy (0/1/2)", "Likert (1-5)",
        "Participant Answer", "Notes",
    ]
    widths = [12, 14, 10, 8, 35, 14, 16, 12, 30, 30]

    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    short_descs = {
        "T1": "Highest avg temp range (season/month)",
        "T2": "Top 3 weather types by frequency",
        "T3": "Year with highest avg wind speed",
        "T4": "Highest avg visibility (humidity > 0.85)",
        "T5": "Top 3 seasons by avg cloud cover",
    }

    row = 2
    for pid in PARTICIPANTS:
        order = SYSTEM_ORDERS[pid]
        order_str = " → ".join(order)
        for sys in order:
            for tid in TASK_IDS:
                values = [
                    pid, order_str, sys, tid,
                    short_descs[tid],
                    "", "", "", "", "",
                ]
                for col, val in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=val)
                    _style_body_cell(cell, wrap=(col >= 5))
                    if col <= 4:
                        cell.alignment = CENTER_ALIGN
                    # Colour-code by system
                    if sys in SYSTEM_FILLS:
                        cell.fill = SYSTEM_FILLS[sys]
                row += 1

    # Freeze header row
    ws.freeze_panes = "A2"


# --- Sheet 4: Post-Study ---

def _create_post_study(wb):
    ws = wb.create_sheet("Post-Study Ranking")
    ws.sheet_properties.tabColor = "8E44AD"

    headers = [
        "Participant",
        "Rank 1 (Best System)",
        "Rank 2",
        "Rank 3 (Worst System)",
        "Reason / Comments",
    ]
    widths = [14, 20, 12, 20, 50]

    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    for row_idx, pid in enumerate(PARTICIPANTS, start=2):
        ws.cell(row=row_idx, column=1, value=pid)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            _style_body_cell(cell, wrap=(col == 5))
            if col <= 4:
                cell.alignment = CENTER_ALIGN
        ws.row_dimensions[row_idx].height = 30

    ws.freeze_panes = "A2"


# --- Main ---

def main():
    wb = openpyxl.Workbook()

    _create_instructions(wb)
    _create_task_reference(wb)
    _create_data_collection(wb)
    _create_post_study(wb)

    os.makedirs(os.path.dirname(OUTPUT_PATH) or ".", exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved evaluation workbook to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
